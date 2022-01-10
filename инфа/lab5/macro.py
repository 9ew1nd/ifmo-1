## author: Рюмин Семён

## подключаем определённые методы из библиотек
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule

## открываем файл с таблицей и "активируем" таблицу
table = load_workbook(filename = 'lab5.xlsx')
ws = table.active

## создаём метод заливки ячейки
redFill = PatternFill(start_color='FF2400',
                      end_color='FF2400')

## создаём курсивный шрифт
itfont = Font(italic = True)

## в цикле проходимся по листам книги и находим нужный
for sheet in table:
    if sheet.title == 'macro':

        ## применяем условное форматирование
        ws.conditional_formatting.add('G4:Y7',
                                          FormulaRule(formula=['G4=1'], fill=redFill))

        ws.conditional_formatting.add('G4:Y7',
                                          FormulaRule(formula=['G4=0'], font = itfont))
        
## сохраняем таблицу
table.save('lab5.xlsx')
